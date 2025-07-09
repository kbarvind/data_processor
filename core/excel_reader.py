"""
Excel reader module for the data processing framework.
Provides utilities for reading, processing, and validating Excel files.
"""

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import ValidationException, handle_exception
from .logging_config import get_logger
from .models import DataRecord

logger = get_logger(__name__)


class ExcelReader:
    """Excel file reader with validation and processing capabilities."""
    
    def __init__(self, file_path: Union[str, Path]):
        """
        Initialize Excel reader.
        
        Args:
            file_path: Path to Excel file
            
        Raises:
            ValidationException: If file doesn't exist or is invalid
        """
        self.file_path = Path(file_path)
        self.workbook: Optional[Workbook] = None
        self.data_frames: Dict[str, pd.DataFrame] = {}
        
        if not self.file_path.exists():
            raise ValidationException(
                f"Excel file not found: {self.file_path}",
                field_name="file_path"
            )
        
        if not self.file_path.suffix.lower() in ['.xlsx', '.xls', '.xlsm']:
            raise ValidationException(
                f"Invalid Excel file format: {self.file_path.suffix}",
                field_name="file_path"
            )
        
        logger.info("ExcelReader initialized", file_path=str(self.file_path))
    
    @handle_exception
    def load_workbook(self) -> Workbook:
        """
        Load Excel workbook.
        
        Returns:
            Loaded workbook
            
        Raises:
            ValidationException: If workbook cannot be loaded
        """
        if self.workbook is None:
            try:
                self.workbook = load_workbook(str(self.file_path), data_only=True)
                logger.info("Workbook loaded successfully", 
                           sheet_count=len(self.workbook.sheetnames))
            except Exception as e:
                raise ValidationException(
                    f"Failed to load Excel workbook: {str(e)}",
                    field_name="file_path",
                    details={"file_path": str(self.file_path)},
                    original_exception=e
                )
        
        return self.workbook
    
    @handle_exception
    def get_sheet_names(self) -> List[str]:
        """
        Get list of sheet names in the workbook.
        
        Returns:
            List of sheet names
        """
        workbook = self.load_workbook()
        return workbook.sheetnames
    
    @handle_exception
    def read_sheet(
        self, 
        sheet_name: Optional[str] = None, 
        header_row: int = 0,
        skip_rows: Optional[List[int]] = None,
        max_rows: Optional[int] = None
    ) -> pd.DataFrame:
        """
        Read data from a specific sheet.
        
        Args:
            sheet_name: Name of the sheet to read (None for first sheet)
            header_row: Row number to use as column headers (0-based)
            skip_rows: List of row numbers to skip
            max_rows: Maximum number of rows to read
            
        Returns:
            DataFrame with sheet data
            
        Raises:
            ValidationException: If sheet cannot be read
        """
        try:
            # Use pandas to read the Excel file
            df = pd.read_excel(
                str(self.file_path),
                sheet_name=sheet_name,
                header=header_row,
                skiprows=skip_rows,
                nrows=max_rows,
                engine='openpyxl'
            )
            
            # If sheet_name is None, pandas returns the first sheet
            if sheet_name is None:
                sheet_names = self.get_sheet_names()
                sheet_name = sheet_names[0] if sheet_names else "Sheet1"
            
            # Clean column names
            df.columns = [str(col).strip() for col in df.columns]
            
            # Store in cache
            self.data_frames[sheet_name] = df
            
            logger.info("Sheet read successfully", 
                       sheet_name=sheet_name,
                       rows=len(df),
                       columns=len(df.columns))
            
            return df
            
        except Exception as e:
            raise ValidationException(
                f"Failed to read Excel sheet '{sheet_name}': {str(e)}",
                field_name="sheet_name",
                details={"sheet_name": sheet_name, "file_path": str(self.file_path)},
                original_exception=e
            )
    
    @handle_exception
    def read_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """
        Read all sheets from the workbook.
        
        Returns:
            Dictionary mapping sheet names to DataFrames
        """
        sheet_names = self.get_sheet_names()
        
        for sheet_name in sheet_names:
            if sheet_name not in self.data_frames:
                self.read_sheet(sheet_name)
        
        logger.info("All sheets read successfully", sheet_count=len(self.data_frames))
        return self.data_frames.copy()
    
    @handle_exception
    def get_sheet_data(self, sheet_name: str) -> pd.DataFrame:
        """
        Get data for a specific sheet (cached if already read).
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            DataFrame with sheet data
        """
        if sheet_name not in self.data_frames:
            self.read_sheet(sheet_name)
        
        return self.data_frames[sheet_name]
    
    @handle_exception
    def validate_columns(
        self, 
        sheet_name: str, 
        required_columns: List[str],
        optional_columns: Optional[List[str]] = None
    ) -> bool:
        """
        Validate that required columns exist in the sheet.
        
        Args:
            sheet_name: Name of the sheet
            required_columns: List of required column names
            optional_columns: List of optional column names
            
        Returns:
            True if validation passes
            
        Raises:
            ValidationException: If required columns are missing
        """
        df = self.get_sheet_data(sheet_name)
        existing_columns = set(df.columns)
        required_set = set(required_columns)
        
        missing_columns = required_set - existing_columns
        
        if missing_columns:
            raise ValidationException(
                f"Missing required columns in sheet '{sheet_name}': {list(missing_columns)}",
                field_name="columns",
                details={
                    "sheet_name": sheet_name,
                    "missing_columns": list(missing_columns),
                    "existing_columns": list(existing_columns)
                }
            )
        
        extra_columns = existing_columns - required_set
        if optional_columns:
            extra_columns = extra_columns - set(optional_columns)
        
        if extra_columns:
            logger.warning("Extra columns found in sheet", 
                          sheet_name=sheet_name,
                          extra_columns=list(extra_columns))
        
        logger.info("Column validation passed", 
                   sheet_name=sheet_name,
                   required_columns=required_columns)
        
        return True
    
    @handle_exception
    def validate_data_types(
        self, 
        sheet_name: str, 
        column_types: Dict[str, str]
    ) -> bool:
        """
        Validate data types of columns.
        
        Args:
            sheet_name: Name of the sheet
            column_types: Dictionary mapping column names to expected types
            
        Returns:
            True if validation passes
            
        Raises:
            ValidationException: If data type validation fails
        """
        df = self.get_sheet_data(sheet_name)
        validation_errors = []
        
        for column, expected_type in column_types.items():
            if column not in df.columns:
                validation_errors.append(f"Column '{column}' not found")
                continue
            
            try:
                if expected_type == 'int':
                    pd.to_numeric(df[column], errors='raise')
                elif expected_type == 'float':
                    pd.to_numeric(df[column], errors='raise')
                elif expected_type == 'datetime':
                    pd.to_datetime(df[column], errors='raise')
                elif expected_type == 'string':
                    df[column].astype(str)
                    
            except Exception as e:
                validation_errors.append(f"Column '{column}' type validation failed: {str(e)}")
        
        if validation_errors:
            raise ValidationException(
                f"Data type validation failed for sheet '{sheet_name}'",
                field_name="data_types",
                validation_errors=validation_errors,
                details={"sheet_name": sheet_name}
            )
        
        logger.info("Data type validation passed", 
                   sheet_name=sheet_name,
                   validated_columns=list(column_types.keys()))
        
        return True
    
    @handle_exception
    def filter_rows(
        self, 
        sheet_name: str, 
        filters: Dict[str, Any]
    ) -> pd.DataFrame:
        """
        Filter rows based on column values.
        
        Args:
            sheet_name: Name of the sheet
            filters: Dictionary of column:value pairs for filtering
            
        Returns:
            Filtered DataFrame
        """
        df = self.get_sheet_data(sheet_name)
        filtered_df = df.copy()
        
        for column, value in filters.items():
            if column not in df.columns:
                raise ValidationException(
                    f"Filter column '{column}' not found in sheet '{sheet_name}'",
                    field_name="filters"
                )
            
            if isinstance(value, list):
                filtered_df = filtered_df[filtered_df[column].isin(value)]
            else:
                filtered_df = filtered_df[filtered_df[column] == value]
        
        logger.info("Rows filtered successfully", 
                   sheet_name=sheet_name,
                   original_rows=len(df),
                   filtered_rows=len(filtered_df),
                   filters=filters)
        
        return filtered_df
    
    @handle_exception
    def convert_to_records(
        self, 
        sheet_name: str, 
        tool_name: str,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[DataRecord]:
        """
        Convert sheet data to DataRecord objects.
        
        Args:
            sheet_name: Name of the sheet
            tool_name: Name of the tool for record metadata
            filters: Optional filters to apply
            
        Returns:
            List of DataRecord objects
        """
        df = self.get_sheet_data(sheet_name)
        
        if filters:
            df = self.filter_rows(sheet_name, filters)
        
        records = []
        for index, row in df.iterrows():
            # Convert row to dictionary, handling NaN values
            row_dict = {}
            for key, value in row.items():
                if pd.isna(value):
                    row_dict[key] = None
                else:
                    row_dict[key] = value
            
            record = DataRecord(
                tool_name=tool_name,
                data=row_dict,
                metadata={
                    "sheet_name": sheet_name,
                    "row_index": index,
                    "source_file": str(self.file_path)
                }
            )
            records.append(record)
        
        logger.info("Records created successfully", 
                   sheet_name=sheet_name,
                   record_count=len(records))
        
        return records
    
    @handle_exception
    def export_to_csv(
        self, 
        sheet_name: str, 
        output_path: Union[str, Path],
        filters: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Export sheet data to CSV file.
        
        Args:
            sheet_name: Name of the sheet
            output_path: Path for output CSV file
            filters: Optional filters to apply
        """
        df = self.get_sheet_data(sheet_name)
        
        if filters:
            df = self.filter_rows(sheet_name, filters)
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        df.to_csv(output_path, index=False)
        
        logger.info("Data exported to CSV", 
                   sheet_name=sheet_name,
                   output_path=str(output_path),
                   rows_exported=len(df))
    
    @handle_exception
    def export_to_json(
        self, 
        sheet_name: str, 
        output_path: Union[str, Path],
        filters: Optional[Dict[str, Any]] = None,
        orient: str = 'records'
    ) -> None:
        """
        Export sheet data to JSON file.
        
        Args:
            sheet_name: Name of the sheet
            output_path: Path for output JSON file
            filters: Optional filters to apply
            orient: JSON orientation ('records', 'index', 'values', etc.)
        """
        df = self.get_sheet_data(sheet_name)
        
        if filters:
            df = self.filter_rows(sheet_name, filters)
        
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Convert to JSON
        json_data = df.to_json(orient=orient, date_format='iso')
        
        with output_path.open('w', encoding='utf-8') as f:
            json.dump(json.loads(json_data), f, indent=2, ensure_ascii=False)
        
        logger.info("Data exported to JSON", 
                   sheet_name=sheet_name,
                   output_path=str(output_path),
                   rows_exported=len(df))
    
    @handle_exception
    def get_summary_stats(self, sheet_name: str) -> Dict[str, Any]:
        """
        Get summary statistics for the sheet.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Dictionary with summary statistics
        """
        df = self.get_sheet_data(sheet_name)
        
        stats = {
            "total_rows": len(df),
            "total_columns": len(df.columns),
            "column_names": list(df.columns),
            "null_counts": df.isnull().sum().to_dict(),
            "data_types": df.dtypes.astype(str).to_dict(),
            "memory_usage": df.memory_usage(deep=True).sum()
        }
        
        # Add numeric statistics for numeric columns
        numeric_columns = df.select_dtypes(include=['int64', 'float64']).columns
        if len(numeric_columns) > 0:
            stats["numeric_stats"] = df[numeric_columns].describe().to_dict()
        
        logger.info("Summary statistics generated", 
                   sheet_name=sheet_name,
                   total_rows=stats["total_rows"],
                   total_columns=stats["total_columns"])
        
        return stats
    
    def close(self):
        """Close the workbook and clear cached data."""
        if self.workbook:
            self.workbook.close()
            self.workbook = None
        
        self.data_frames.clear()
        logger.info("ExcelReader closed")
    
    def __enter__(self):
        """Context manager entry."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()


class ExcelWriter:
    """Excel file writer for creating and updating Excel files."""
    
    def __init__(self, file_path: Union[str, Path]):
        """
        Initialize Excel writer.
        
        Args:
            file_path: Path for output Excel file
        """
        self.file_path = Path(file_path)
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        logger.info("ExcelWriter initialized", file_path=str(self.file_path))
    
    @handle_exception
    def create_sheet(self, sheet_name: str, data: pd.DataFrame) -> None:
        """
        Create a new sheet with data.
        
        Args:
            sheet_name: Name of the sheet
            data: DataFrame to write to sheet
        """
        # Create new worksheet
        ws = self.workbook.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, column in enumerate(data.columns, 1):
            ws.cell(row=1, column=col_idx, value=column)
        
        # Write data
        for row_idx, row in enumerate(data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        logger.info("Sheet created successfully", 
                   sheet_name=sheet_name,
                   rows=len(data),
                   columns=len(data.columns))
    
    @handle_exception
    def save(self) -> None:
        """Save the workbook to file."""
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(str(self.file_path))
        
        logger.info("Workbook saved", file_path=str(self.file_path))
    
    def close(self):
        """Close the workbook."""
        self.workbook.close()
        logger.info("ExcelWriter closed")
    
    def __enter__(self):
        """Context manager entry."""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit."""
        self.close()


def read_excel_file(
    file_path: Union[str, Path],
    sheet_name: Optional[str] = None,
    validation_config: Optional[Dict[str, Any]] = None
) -> pd.DataFrame:
    """
    Convenience function to read Excel file with optional validation.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to read (None for first sheet)
        validation_config: Optional validation configuration
        
    Returns:
        DataFrame with Excel data
    """
    with ExcelReader(file_path) as reader:
        df = reader.read_sheet(sheet_name)
        
        if validation_config:
            # Apply validation if configured
            if 'required_columns' in validation_config:
                reader.validate_columns(
                    sheet_name or reader.get_sheet_names()[0],
                    validation_config['required_columns'],
                    validation_config.get('optional_columns')
                )
            
            if 'column_types' in validation_config:
                reader.validate_data_types(
                    sheet_name or reader.get_sheet_names()[0],
                    validation_config['column_types']
                )
        
        return df 